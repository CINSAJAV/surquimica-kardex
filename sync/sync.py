"""
sync.py — Kardex Surquimica
Lee todos los .xlsx desde la carpeta de Google Drive local y genera JSON
para el frontend. Ejecutar antes de cada deploy o cuando cambian los Excel.

Uso:
    python sync.py
    python sync.py --drive "G:\\ruta\\alternativa"

Salida:
    ../web/public/data/index.json  — catálogo de productos
    ../web/public/data/K-N.json   — movimientos por producto
"""

import json
import os
import re
import sys
from datetime import datetime, date
from pathlib import Path

import openpyxl

# ─── CONFIGURACIÓN ────────────────────────────────────────────────────────────

DRIVE_PATH      = Path(r"G:\Mi unidad\Surquimica\Kardex")
BODEGA_PATH     = Path(r"G:\Mi unidad\Surquimica\INVENTARIO BODEGAS SQ 29052026.xlsx")
PRECIOS_GLOB    = r"G:\Mi unidad\Surquimica\Clientes_K*.xlsx"
DESPACHOS_PATH  = Path(r"G:\Mi unidad\Surquimica\Consolidado_Despachos_Clientes_2026.xlsx")
OUTPUT_DIR      = Path(__file__).parent.parent / "public" / "data"

# Patrones de texto en detalle que indican que NO es nombre de cliente
SKIP_PATTERNS = re.compile(
    r"\b(OC|GD|F-|K\d|PARCIAL|AGENCIA|CONTRATO|F\s*\d{4,})\b", re.IGNORECASE
)

# Palabras/frases de proveedores conocidos (para clasificación compra)
PROVEEDOR_KEYWORDS = ["BRENNTAG", "JEBSEN", "QUIMICA DEL PACIFICO", "ERCROS",
                      "UNIVAR", "SOLQUIM", "TRANSA"]


# ─── PARSEO DE FECHA ──────────────────────────────────────────────────────────

def parse_fecha(val) -> str | None:
    """Normaliza fecha a ISO YYYY-MM-DD. Acepta datetime, date, DD-MM-YYYY, M/D/YYYY."""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s or s.lower() in ("none", "fecha", "n/a"):
        return None
    # DD-MM-YYYY
    m = re.match(r"^(\d{1,2})-(\d{1,2})-(\d{4})$", s)
    if m:
        d, mo, y = m.groups()
        return f"{y}-{mo.zfill(2)}-{d.zfill(2)}"
    # M/D/YYYY  o  MM/DD/YYYY
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)
    if m:
        mo, d, y = m.groups()
        return f"{y}-{mo.zfill(2)}-{d.zfill(2)}"
    # YYYY-MM-DD (ya normalizado)
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", s)
    if m:
        return s
    return None


# ─── EXTRACCIÓN DE CLIENTE ────────────────────────────────────────────────────

# Sufijos legales a normalizar
_SUFIJOS = re.compile(
    r"\s+(S\.A\.|LTDA\.?|E\.I\.R\.L\.?|SPA\.?|S\.p\.A\.?|AG\.?|GMBH\.?)\.?$",
    re.IGNORECASE,
)
# Tokens que no son parte del nombre cliente
_TOKENS_OC = re.compile(
    r"\b(OC|GD-?|F-|K\d+|PARCIAL|AGENCIA\s*N[°º]?\s*\d+|N[°º]?\s*\d+|"
    r"CONTRATO\s+PENDIENTE|APA|CLASE\s+\d+)\b.*",
    re.IGNORECASE,
)
# Referencia de contrato tipo "C-4313" al inicio o al final
_CONTRATO_INICIO = re.compile(r"^[A-Z]-\d+\s+", re.IGNORECASE)
_CONTRATO_FINAL  = re.compile(r"\s+[A-Z]-\d+.*$", re.IGNORECASE)
# Tokens de despacho/ruta: "E.P.", "EP.", "DESPACHO N" (consume la puntuación final también)
_DESPACHO_TOKENS = re.compile(r"\b(?:E\.P\.|E\.P|EP\.?|DESPACHO\s*\d*)\s*", re.IGNORECASE)


def extraer_cliente(detalle: str) -> str:
    if not detalle:
        return "DESCONOCIDO"
    nombre = detalle.strip()
    # Quitar prefijos de documento al inicio
    # "OC XXXXXXX NOMBRE"
    nombre = re.sub(r"^OC\s+\d+\s+", "", nombre, flags=re.IGNORECASE)
    # "FACTURA[FT]XXXX NOMBRE"
    nombre = re.sub(r"^FACTURA\s*\[FT\]\s*\d+\s+", "", nombre, flags=re.IGNORECASE)
    # "GUÍA[ST]XXXX NOMBRE" / "GUÍA[EA]XXXX NOMBRE" (incluyendo variantes de encoding del acento)
    nombre = re.sub(r"^GU.{0,2}A\s*\[(?:ST|EA)\]\s*\d+\s+", "", nombre, flags=re.IGNORECASE)
    # Quitar sufijos que no son parte del nombre (primera pasada, antes de _TOKENS_OC)
    # "... COT XXXX" o "... COT-XXXX"
    nombre = re.sub(r"\s+COT[-\s]*\d+.*$", "", nombre, flags=re.IGNORECASE)
    # "... XXXXXXXXXX-XXXX" (número tipo RUT/correlativo al final)
    nombre = re.sub(r"\s+\d{8,}-\d+\s*$", "", nombre)
    # Quitar referencia contrato al inicio (ej: "C-4313 E.P. AGUAS DE...")
    nombre = _CONTRATO_INICIO.sub("", nombre)
    # Quitar referencia contrato al final (ej: "AGUAS DE ANTOFAGASTA S.A.  C-4313 DESPACHO 2 EP")
    nombre = _CONTRATO_FINAL.sub("", nombre)
    # Quitar tokens de despacho/ruta (E.P., EP., DESPACHO N)
    nombre = _DESPACHO_TOKENS.sub("", nombre)
    # Limpiar puntuación sobrante al inicio y fin
    nombre = nombre.strip().strip(",.-").strip()
    # Quitar número de OC/factura al inicio (ej: "8656 EMP.SERV..." → "EMP.SERV...")
    nombre = re.sub(r"^\d+\s+", "", nombre)
    # Quitar tokens de OC / GD / F- etc. y todo lo que sigue
    nombre = _TOKENS_OC.sub("", nombre).strip().rstrip(",.-")
    # Quitar sufijos legales para normalizar
    nombre = _SUFIJOS.sub("", nombre).strip()
    # Segunda pasada de sufijos (pueden quedar expuestos tras _TOKENS_OC)
    nombre = re.sub(r"\s+P\d{4,}\s*$", "", nombre, flags=re.IGNORECASE).strip()
    nombre = re.sub(r"\s+COT[-\s]*\d+.*$", "", nombre, flags=re.IGNORECASE).strip()
    return nombre.upper() if nombre else detalle.upper()


def es_proveedor(detalle: str) -> bool:
    d = detalle.upper()
    return any(k in d for k in PROVEEDOR_KEYWORDS)


# ─── PARSEO DE UN BLOQUE KARDEX ───────────────────────────────────────────────

def parse_bloque(rows: list, start: int, end: int, archivo: str) -> dict | None:
    """
    Parsea un bloque de filas [start, end) que contiene un Kardex individual.
    rows: lista de tuplas (valores de celdas, 0-indexed columnas).
    Estructura esperada dentro del bloque:
      +0: "Tarjeta de Existencia Financiera"
      +1: "KARDEX N - NOMBRE" o "NOMBRE  Kardex N"
      +2: unidad (KILOS / LITROS)
      +5: encabezado Fecha | Nº Factura | Detalle | ...
      +6: SALDO INICIAL
      +7+: movimientos
    """
    if end - start < 6:
        return None

    # Fila 1 del bloque: nombre y código
    fila_nombre = rows[start + 1]
    nombre_raw = str(fila_nombre[1] or "").strip() if len(fila_nombre) > 1 else ""

    # Formato A: "KARDEX 36 - SULFATO DE ALUMINIO FINO"
    m = re.match(r"KARDEX\s+([\d\-]+)\s*[-–]\s*(.+)", nombre_raw, re.IGNORECASE)
    if not m:
        m = re.match(r"KARDEX\s+([\d\-]+)\s+(.+)", nombre_raw, re.IGNORECASE)
    if m:
        codigo_raw = m.group(1).strip()
        nombre = m.group(2).strip()
    else:
        # Formato B: "NOMBRE PRODUCTO  Kardex N"
        m2 = re.search(r"Kardex\s+([\d\-]+)\s*$", nombre_raw, re.IGNORECASE)
        if m2:
            codigo_raw = m2.group(1).strip()
            nombre = nombre_raw[:m2.start()].strip()
        else:
            # Fallback: código desde nombre del archivo "K N NOMBRE - 2026.xlsx"
            mf = re.match(r"K\s+([\d\-]+)\s+", archivo)
            codigo_raw = mf.group(1).strip() if mf else "?"
            nombre = nombre_raw or archivo.replace("- 2026.xlsx", "").strip()

    codigo = f"K-{codigo_raw}"

    # Fila 2: unidad
    fila_unidad = rows[start + 2]
    unidad_raw = str(fila_unidad[1] or "").upper() if len(fila_unidad) > 1 else ""
    if "KILO" in unidad_raw:
        unidad = "kg"
    elif "LITRO" in unidad_raw or "LTS" in unidad_raw or "LT" in unidad_raw:
        unidad = "lt"
    else:
        unidad = "un"

    # Fila SALDO INICIAL (generalmente start+6, pero buscar por "SALDO INICIAL")
    saldo_inicial_row = None
    header_row = None
    for i in range(start + 3, min(start + 10, end)):
        row = rows[i]
        detalle_cell = str(row[3] or "").strip() if len(row) > 3 else ""
        fecha_cell = str(row[1] or "").strip() if len(row) > 1 else ""
        if "SALDO INICIAL" in detalle_cell.upper():
            saldo_inicial_row = i
        if "FECHA" in fecha_cell.upper() or "Fecha" in str(row[1] or ""):
            header_row = i

    # Si no encontramos saldo inicial, tomar start+6 por defecto
    if saldo_inicial_row is None:
        saldo_inicial_row = start + 6
    mov_start = saldo_inicial_row + 1

    # Leer saldo inicial
    si = rows[saldo_inicial_row] if saldo_inicial_row < end else [None] * 12
    stock_inicial = _num(si[4] if len(si) > 4 else None) or 0
    saldo_ini_pesos = _num(si[9] if len(si) > 9 else None) or 0
    pmp_inicial = _num(si[10] if len(si) > 10 else None) or 0

    # Movimientos
    movimientos = []
    total_entradas = 0
    total_salidas = 0
    total_haber = 0

    for i in range(mov_start, end):
        row = rows[i]
        if len(row) < 7:
            continue

        fecha_val = parse_fecha(row[1] if len(row) > 1 else None)
        detalle = str(row[3] or "").strip() if len(row) > 3 else ""
        entrada = _num(row[4] if len(row) > 4 else None) or 0
        salida = _num(row[5] if len(row) > 5 else None) or 0
        saldo_kg = _num(row[6] if len(row) > 6 else None)
        haber = _num(row[8] if len(row) > 8 else None) or 0
        saldo_p = _num(row[9] if len(row) > 9 else None)
        pmp = _num(row[10] if len(row) > 10 else None) or 0

        # Saltar fila TOTALES (por texto o por tener ambas columnas > 0)
        if "TOTAL" in detalle.upper():
            break
        if entrada > 0 and salida > 0:
            # Fila de totales sin etiqueta "TOTAL" — ignorar y parar
            break
        if not fecha_val and entrada == 0 and salida == 0:
            continue
        if entrada == 0 and salida == 0:
            continue

        tipo = "Compra" if entrada > 0 else "Venta"
        cliente = extraer_cliente(detalle)

        mov = {
            "fecha": fecha_val or "",
            "detalle": detalle,
            "cliente": cliente,
            "tipo": tipo,
            "entrada": entrada,
            "salida": salida,
            "saldo_kg": saldo_kg,
            "haber": haber,
            "saldo_p": saldo_p,
            "pmp": round(pmp, 4) if pmp else 0,
        }
        movimientos.append(mov)
        total_entradas += entrada
        total_salidas += salida
        total_haber += haber

    if not movimientos and stock_inicial == 0:
        return None  # bloque vacío

    # Stock y PMP actuales: último movimiento con saldo_kg no nulo
    stock_actual = stock_inicial
    pmp_actual = pmp_inicial
    saldo_actual_pesos = saldo_ini_pesos
    for mv in reversed(movimientos):
        if mv["saldo_kg"] is not None:
            stock_actual = mv["saldo_kg"]
            pmp_actual = mv["pmp"]
            saldo_actual_pesos = mv["saldo_p"] or 0
            break

    return {
        "codigo": codigo,
        "nombre": nombre,
        "unidad": unidad,
        "archivo": archivo,
        "stockInicial": stock_inicial,
        "saldoInicialPesos": saldo_ini_pesos,
        "pmpInicial": pmp_inicial,
        "stockActual": stock_actual,
        "pmpActual": round(pmp_actual, 4),
        "saldoActualPesos": saldo_actual_pesos,
        "totalEntradas": total_entradas,
        "totalSalidas": total_salidas,
        "totalHaber": total_haber,
        "movimientos": movimientos,
    }


def _num(val) -> float | None:
    """Convierte celda a float, devuelve None si no es numérico."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


# ─── PARSEO DE UN ARCHIVO XLSX ────────────────────────────────────────────────

def parse_archivo(path: Path) -> list[dict]:
    """Devuelve lista de productos encontrados en el archivo."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        print(f"  ERROR abriendo {path.name}: {e}")
        return []

    # Detectar inicios de bloques KARDEX
    bloque_starts = []
    for i, row in enumerate(rows):
        cell = str(row[1] or "").strip() if len(row) > 1 else ""
        if re.match(r"KARDEX\s+\d", cell, re.IGNORECASE):
            bloque_starts.append(i - 1)  # incluir fila 0 del bloque (título)

    if not bloque_starts:
        # Archivo de un solo producto: buscar desde fila 0
        bloque_starts = [0]

    productos = []
    for idx, start in enumerate(bloque_starts):
        end = bloque_starts[idx + 1] if idx + 1 < len(bloque_starts) else len(rows)
        prod = parse_bloque(rows, start, end, path.name)
        if prod:
            productos.append(prod)

    return productos


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    drive_path = DRIVE_PATH
    if "--drive" in sys.argv:
        i = sys.argv.index("--drive")
        drive_path = Path(sys.argv[i + 1])

    if not drive_path.exists():
        print(f"ERROR: No se encontró la carpeta Drive: {drive_path}")
        sys.exit(1)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    archivos = sorted(drive_path.glob("K *.xlsx"))
    print(f"Encontrados {len(archivos)} archivos en {drive_path}\n")

    todos_productos = []
    prods_full = []   # lista completa con movimientos (para empresas.json)
    errores = []

    for arch in archivos:
        print(f"Procesando: {arch.name}")
        prods = parse_archivo(arch)
        # Si el nombre de archivo tiene variante decimal (K 8.1, K 8.2), forzar codigo
        mv = re.match(r"K\s+(\d+\.\d+)\s", arch.name)
        if mv and prods:
            codigo_forzado = f"K-{mv.group(1).replace('.', '-')}"
            for pr in prods:
                pr["codigo"] = codigo_forzado
        if not prods:
            errores.append(arch.name)
            print(f"  → Sin productos extraídos")
        for p in prods:
            prods_full.append(p)
            print(f"  OK {p['codigo']} {p['nombre']} | "
                  f"stock={p['stockActual']:,.0f} {p['unidad']} | "
                  f"movs={len(p['movimientos'])}")
            # Guardar JSON individual (puntos → guiones para slugs limpios)
            slug = re.sub(r"[^\w\-]", "", p["codigo"].replace(" ", "-").replace(".", "-"))
            out_file = OUTPUT_DIR / f"{slug}.json"
            with open(out_file, "w", encoding="utf-8") as f:
                json.dump(p, f, ensure_ascii=False, indent=2, default=str)
            todos_productos.append({
                "codigo": p["codigo"],
                "nombre": p["nombre"],
                "unidad": p["unidad"],
                "archivo": p["archivo"],
                "stockActual": p["stockActual"],
                "pmpActual": p["pmpActual"],
                "saldoActualPesos": p["saldoActualPesos"],
                "totalEntradas": p["totalEntradas"],
                "totalSalidas": p["totalSalidas"],
                "totalHaber": p["totalHaber"],
                "nMovimientos": len(p["movimientos"]),
                "slug": slug,
            })

    # Índice general
    todos_productos.sort(key=lambda p: p["codigo"])
    index = {
        "generado": datetime.now().isoformat(timespec="seconds"),
        "total": len(todos_productos),
        "productos": todos_productos,
    }
    with open(OUTPUT_DIR / "index.json", "w", encoding="utf-8") as f:
        json.dump(index, f, ensure_ascii=False, indent=2, default=str)

    # ── Agregado por empresa ──────────────────────────────────────────────────
    empresas: dict[str, dict] = {}

    for p in prods_full:
        for mv in p["movimientos"]:
            if mv["tipo"] != "Venta":
                continue
            nombre = mv["cliente"]
            if not nombre or nombre == "DESCONOCIDO":
                continue

            if nombre not in empresas:
                empresas[nombre] = {
                    "nombre": nombre,
                    "totalHaber": 0,
                    "totalCosto": 0,
                    "totalVol": 0,
                    "nMovimientos": 0,
                    "ultimaFecha": "",
                    "primeraFecha": "",
                    "productos": {},
                }
            e = empresas[nombre]
            e["totalHaber"]   += mv["haber"] or 0
            e["totalCosto"]   += (mv["salida"] or 0) * (mv["pmp"] or 0)
            e["totalVol"]     += mv["salida"] or 0
            e["nMovimientos"] += 1
            if mv["fecha"]:
                if not e["ultimaFecha"] or mv["fecha"] > e["ultimaFecha"]:
                    e["ultimaFecha"] = mv["fecha"]
                if not e["primeraFecha"] or mv["fecha"] < e["primeraFecha"]:
                    e["primeraFecha"] = mv["fecha"]

            slug = re.sub(r"[^\w\-]", "", p["codigo"].replace(" ", "-").replace(".", "-"))
            if slug not in e["productos"]:
                e["productos"][slug] = {
                    "codigo": p["codigo"],
                    "slug": slug,
                    "nombre": p["nombre"],
                    "unidad": p["unidad"],
                    "vol": 0,
                    "monto": 0,
                    "costo": 0,
                    "nMovimientos": 0,
                    "ultimaFecha": "",
                    "primeraFecha": "",
                }
            ep = e["productos"][slug]
            ep["vol"]          += mv["salida"] or 0
            ep["monto"]        += mv["haber"] or 0
            ep["costo"]        += (mv["salida"] or 0) * (mv["pmp"] or 0)
            ep["nMovimientos"] += 1
            if mv["fecha"]:
                if not ep["ultimaFecha"] or mv["fecha"] > ep["ultimaFecha"]:
                    ep["ultimaFecha"] = mv["fecha"]
                if not ep["primeraFecha"] or mv["fecha"] < ep["primeraFecha"]:
                    ep["primeraFecha"] = mv["fecha"]

    # Convertir productos dict → lista ordenada por monto desc
    empresas_list = []
    for e in empresas.values():
        e["productos"] = sorted(e["productos"].values(), key=lambda x: x["monto"], reverse=True)
        e["nProductos"] = len(e["productos"])
        empresas_list.append(e)

    empresas_list.sort(key=lambda x: x["totalHaber"], reverse=True)

    empresas_json = {
        "generado": datetime.now().isoformat(timespec="seconds"),
        "total": len(empresas_list),
        "empresas": empresas_list,
    }
    with open(OUTPUT_DIR / "empresas.json", "w", encoding="utf-8") as f:
        json.dump(empresas_json, f, ensure_ascii=False, indent=2, default=str)

    print(f"OK {len(empresas_list)} empresas -> empresas.json")
    # ── Inventario bodega ─────────────────────────────────────────────────────
    if BODEGA_PATH.exists():
        bodega = parse_bodega(BODEGA_PATH)
        bodega_out = {
            "generado": datetime.now().isoformat(timespec="seconds"),
            "archivo":  BODEGA_PATH.name,
            "productos": bodega,
        }
        with open(OUTPUT_DIR / "bodega.json", "w", encoding="utf-8") as f:
            json.dump(bodega_out, f, ensure_ascii=False, indent=2, default=str)
        print(f"OK {len(bodega)} productos bodega -> bodega.json")
    else:
        print(f"WARN bodega no encontrada: {BODEGA_PATH}")

    # ── Precios por cliente ───────────────────────────────────────────────────
    precios = parse_precios()
    if precios:
        precios_json = {
            "generado": datetime.now().isoformat(timespec="seconds"),
            "productos": precios,
        }
        with open(OUTPUT_DIR / "precios.json", "w", encoding="utf-8") as f:
            json.dump(precios_json, f, ensure_ascii=False, indent=2, default=str)
        total_cl = sum(len(v) for v in precios.values())
        print(f"OK {len(precios)} productos / {total_cl} registros precios -> precios.json")

    # ── Costos de despacho ───────────────────────────────────────────────────
    despachos = parse_despachos()
    if despachos:
        despachos_json = {
            "generado": datetime.now().isoformat(timespec="seconds"),
            "productos": despachos,
        }
        with open(OUTPUT_DIR / "despachos.json", "w", encoding="utf-8") as f:
            json.dump(despachos_json, f, ensure_ascii=False, indent=2, default=str)
        total_d = sum(len(v) for v in despachos.values())
        print(f"OK {len(despachos)} productos / {total_d} registros despacho -> despachos.json")

    print(f"\nOK {len(todos_productos)} productos -> {OUTPUT_DIR}")
    if errores:
        print(f"WARN Con errores: {errores}")


# ─── PARSEO INVENTARIO BODEGA ─────────────────────────────────────────────────

def parse_bodega(path: Path) -> dict:
    """
    Lee INVENTARIO BODEGAS SQ .xlsx y devuelve dict indexado por slug de producto.
    Columnas (0-based):
      0: código KARDEX  1: nombre
      2: softland_abr   3: fisico_abr   4: diff_abr   5: obs_abr
      6: softland_may   7: fisico_may   8: diff_may   9: obs_may
    """
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
    except Exception as e:
        print(f"  WARN bodega: {e}")
        return {}

    resultado = {}
    # Intentar extraer fecha de encabezado (fila 1)
    header = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0] if False else None

    for row in rows:
        kod_raw = str(row[0] or "").strip()
        nombre  = str(row[1] or "").strip()
        if not kod_raw or not nombre:
            continue

        # Ignorar remanentes (código termina en R o contiene REMANENTE)
        if re.search(r"\bR\b", kod_raw, re.IGNORECASE) or "REMANENTE" in nombre.upper():
            continue

        # Extraer número del código: "KARDEX1 $" → "1", "KARDEX 94" → "94"
        m = re.search(r"KARDEX\s*(\d+)", kod_raw, re.IGNORECASE)
        if not m:
            continue
        slug = f"K-{m.group(1)}"

        def to_num(v):
            if v is None: return None
            try: return float(str(v).replace(" ","").replace("-","0") if str(v).strip() in ("-"," -  ","  - ") else v)
            except: return None

        softland_abr = to_num(row[2])
        fisico_abr   = to_num(row[3])
        diff_abr     = to_num(row[4])
        obs_abr      = str(row[5]).strip() if row[5] else None
        softland_may = to_num(row[6])
        fisico_may   = to_num(row[7])
        diff_may     = to_num(row[8])
        obs_may      = str(row[9]).strip() if row[9] else None

        resultado[slug] = {
            "nombre":  nombre,
            "abril": {
                "fecha":    "2026-04-30",
                "softland": softland_abr,
                "fisico":   fisico_abr,
                "diferencia": diff_abr,
                "obs":      obs_abr,
            },
            "mayo": {
                "fecha":    "2026-05-29",
                "softland": softland_may,
                "fisico":   fisico_may,
                "diferencia": diff_may,
                "obs":      obs_may,
            },
        }

    return resultado


# ─── PARSEO ARCHIVOS DE PRECIOS POR CLIENTE ──────────────────────────────────

def parse_precios() -> dict:
    """
    Lee todos los Clientes_K*.xlsx de G:\\Mi unidad\\Surquimica\\ y devuelve dict:
    { "K-1": { "CLIENTE A": { "cantidad": x, "pmp_pond": x, "precio_venta": x, "margen": x } } }
    """
    import glob as _glob
    archivos = sorted(_glob.glob(PRECIOS_GLOB))
    if not archivos:
        print(f"  WARN no se encontraron archivos de precios en: {PRECIOS_GLOB}")
        return {}

    resultado = {}

    for ruta in archivos:
        path = Path(ruta)
        # Extraer código del nombre: Clientes_K1_... → K-1, Clientes_K8L_... → K-8-2
        # Soporta: K8 → K-8, K8.1 → K-8-1, K8L → K-8-2 (L = litros, segunda variante)
        m = re.match(r"Clientes_K(\d+(?:\.\d+)?)(L?)_", path.name, re.IGNORECASE)
        if not m:
            continue
        num_part = m.group(1).replace(".", "-")
        l_suffix = m.group(2).upper()
        if l_suffix == "L":
            # Variante litros: busca K-{num}-2 si existe, si no K-{num}L
            slug = f"K-{num_part}-2"
        else:
            slug = f"K-{num_part}"

        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
        except Exception as e:
            print(f"  ERROR leyendo {path.name}: {e}")
            continue

        # Buscar fila de encabezado: "Cliente" como celda propia (no en oración larga)
        # y que también tenga "Precio" o "Margen" en la misma fila
        header_idx = None
        for i, row in enumerate(rows):
            vals = [str(c or "").strip().lower() for c in row]
            has_cliente = any(v == "cliente" or (len(v) < 15 and "cliente" in v) for v in vals)
            has_precio = any("precio" in v or "margen" in v for v in vals)
            if has_cliente and has_precio:
                header_idx = i
                break

        if header_idx is None:
            continue

        # Mapear columnas: N°, Cliente, Cantidad Total, PMP Pond., PMP Último, Precio Venta, Margen
        hrow = rows[header_idx]
        col_map = {}
        for ci, cell in enumerate(hrow):
            t = str(cell or "").strip().lower()
            if "cliente" in t:
                col_map["cliente"] = ci
            elif "cantidad" in t:
                col_map["cantidad"] = ci
            elif "pond" in t:
                col_map["pmp_pond"] = ci
            elif "ltimo" in t or "ultimo" in t:
                col_map["pmp_ultimo"] = ci
            elif "precio" in t:
                col_map["precio_venta"] = ci
            elif "margen" in t:
                col_map["margen"] = ci

        if "cliente" not in col_map or "precio_venta" not in col_map:
            print(f"  WARN {path.name}: no se encontraron columnas esperadas")
            continue

        clientes = {}
        for row in rows[header_idx + 1:]:
            if not row or all(c is None for c in row):
                continue
            num_val = str(row[0] or "").strip()
            # Saltar filas de sub-entrada (ej: "1b", "3b") — mismas condiciones, diferente período
            if re.match(r"^\d+[a-z]+$", num_val, re.IGNORECASE):
                continue
            # Parar en fila TOTAL
            cliente_raw = str(row[col_map["cliente"]] or "").strip()
            if not cliente_raw or "TOTAL" in cliente_raw.upper():
                continue

            precio_raw = row[col_map["precio_venta"]] if "precio_venta" in col_map else None
            margen_raw = row[col_map.get("margen", -1)] if "margen" in col_map else None
            cantidad_raw = row[col_map.get("cantidad", -1)] if "cantidad" in col_map else None
            pmp_pond_raw = row[col_map.get("pmp_pond", -1)] if "pmp_pond" in col_map else None

            precio = _num(precio_raw)
            margen = _num(margen_raw)
            cantidad = _num(cantidad_raw)
            pmp_pond = _num(pmp_pond_raw)
            pmp_ultimo_raw = row[col_map.get("pmp_ultimo", -1)] if "pmp_ultimo" in col_map else None
            pmp_ultimo = _num(pmp_ultimo_raw)

            # Normalizar nombre cliente (quitar sufijos legales)
            nombre = _SUFIJOS.sub("", cliente_raw).strip().upper()

            clientes[nombre] = {
                "cantidad":    round(cantidad, 2)   if cantidad   is not None else None,
                "pmp_pond":    round(pmp_pond, 2)   if pmp_pond   is not None else None,
                "pmp_ultimo":  round(pmp_ultimo, 2) if pmp_ultimo is not None else None,
                "precio_venta":round(precio, 2)     if precio     is not None else None,
                "margen":      round(margen, 2)     if margen     is not None else None,
            }

        if clientes:
            resultado[slug] = clientes
            print(f"  OK precios {slug}: {len(clientes)} clientes")

    return resultado


# ─── PARSEO COSTOS DE DESPACHO ───────────────────────────────────────────────

def parse_despachos() -> dict:
    """
    Lee Consolidado_Despachos_Clientes_2026.xlsx y devuelve dict:
    { "K-4": { "SURALIS S.A.": 112.5, ... }, ... }
    Valor = costo promedio ponderado de transporte en $/kg.
    """
    from openpyxl import load_workbook as _lw

    if not DESPACHOS_PATH.exists():
        print(f"  [despachos] No encontrado: {DESPACHOS_PATH}")
        return {}

    wb  = _lw(DESPACHOS_PATH, data_only=True)
    ws  = wb.active

    # Encontrar fila de encabezado (tiene "Kardex" y "Valor por Kilo")
    header_idx = None
    col_map    = {}
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        row_str = [str(c).strip() if c is not None else "" for c in row]
        joined  = " ".join(row_str).lower()
        if "kardex" in joined and "valor por kilo" in joined:
            header_idx = i
            for j, h in enumerate(row_str):
                hl = h.lower()
                if hl == "cliente":
                    col_map["cliente"] = j
                elif hl == "kardex":
                    col_map["kardex"]  = j
                elif hl == "cantidad (kg)":
                    col_map["cantidad"] = j
                elif hl == "valor por kilo":
                    col_map["costo_kg"] = j
            break

    if header_idx is None or len(col_map) < 4:
        print(f"  [despachos] Header no encontrado o columnas incompletas: {col_map}")
        return {}

    # Acumular (slug, cliente) → {kg, costo_total}
    acum: dict = {}
    for row in ws.iter_rows(min_row=header_idx + 1, values_only=True):
        try:
            cliente  = str(row[col_map["cliente"]] or "").strip().upper()
            kardex   = int(row[col_map["kardex"]])
            cantidad = float(row[col_map["cantidad"]])
            costo_kg = float(row[col_map["costo_kg"]])
        except (TypeError, ValueError):
            continue

        if not cliente or cantidad <= 0 or costo_kg <= 0:
            continue

        slug = f"K-{kardex}"
        acum.setdefault(slug, {}).setdefault(cliente, {"kg": 0.0, "costo_total": 0.0})
        acum[slug][cliente]["kg"]          += cantidad
        acum[slug][cliente]["costo_total"] += costo_kg * cantidad

    # Calcular promedio ponderado
    result: dict = {}
    for slug, clientes in acum.items():
        result[slug] = {
            cliente: round(d["costo_total"] / d["kg"], 2)
            for cliente, d in clientes.items()
            if d["kg"] > 0
        }
        print(f"  OK despachos {slug}: {len(result[slug])} clientes")

    return result


if __name__ == "__main__":
    main()
